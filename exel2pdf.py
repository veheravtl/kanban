#!/usr/bin/env python3

from pathlib import Path
import argparse
import os
import shutil
import subprocess
import tempfile
import sys
import zipfile

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.properties import PageSetupProperties
    OPENPYXL_IMPORT_ERROR = None
except ModuleNotFoundError as exc:
    load_workbook = None
    get_column_letter = None
    PageSetupProperties = None
    OPENPYXL_IMPORT_ERROR = exc


class ValidationError(Exception):
    pass


SUPPORTED_EXCEL_EXTENSIONS = {".xls", ".xlsx", ".xlsm"}
SUPPORTED_WORD_EXTENSIONS = {".docx"}
SUPPORTED_EXTENSIONS = SUPPORTED_EXCEL_EXTENSIONS | SUPPORTED_WORD_EXTENSIONS
SOFFICE_STARTUP_FLAGS = [
    "--headless",
    "--nologo",
    "--nodefault",
    "--nolockcheck",
    "--norestore",
    "--invisible",
]


def find_soffice() -> str:
    path = shutil.which("soffice")
    if path:
        return path

    path = shutil.which("libreoffice")
    if path:
        return path

    raise ValidationError("LibreOffice is not installed or not in PATH")


def validate_input_file(src: Path) -> None:
    if not src.exists():
        raise ValidationError(f"Input file not found: {src}")

    if not src.is_file():
        raise ValidationError(f"Input path is not a file: {src}")

    if src.suffix.lower() not in SUPPORTED_EXTENSIONS:
        allowed = ", ".join(sorted(SUPPORTED_EXTENSIONS))
        raise ValidationError(f"Only {allowed} are supported")

    if src.suffix.lower() == ".docx":
        validate_docx_file(src)


def validate_docx_file(src: Path) -> None:
    try:
        with zipfile.ZipFile(src) as zf:
            if "word/document.xml" not in zf.namelist():
                raise ValidationError("Input .docx file is invalid or damaged")
    except zipfile.BadZipFile:
        raise ValidationError("Input .docx file is invalid or damaged")


def is_excel_file(src: Path) -> bool:
    return src.suffix.lower() in SUPPORTED_EXCEL_EXTENSIONS


def get_visible_sheets(workbook):
    return [ws for ws in workbook.worksheets if ws.sheet_state == "visible"]


def is_effectively_empty(ws) -> bool:
    for _ in iter_cells_with_content(ws):
        return False
    return True


def has_cell_content(cell) -> bool:
    value = cell.value
    if value is None:
        return False
    if isinstance(value, str) and value.strip() == "":
        return False
    return True


def iter_cells_with_content(ws):
    cells = getattr(ws, "_cells", None)
    if isinstance(cells, dict):
        for cell in cells.values():
            if has_cell_content(cell):
                yield cell
        return

    for row in ws.iter_rows():
        for cell in row:
            if has_cell_content(cell):
                yield cell


def find_last_used_cell(ws):
    max_row = 0
    max_col = 0

    for cell in iter_cells_with_content(ws):
        if cell.row > max_row:
            max_row = cell.row
        if cell.column > max_col:
            max_col = cell.column

    # Merged cells keep value only in the top-left cell, but printed width
    # must include the whole merged range.
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, range_max_col, range_max_row = merged_range.bounds
        anchor_cell = ws.cell(row=min_row, column=min_col)
        if has_cell_content(anchor_cell):
            if range_max_row > max_row:
                max_row = range_max_row
            if range_max_col > max_col:
                max_col = range_max_col

    if max_row == 0 or max_col == 0:
        raise ValidationError("Worksheet is empty")

    return max_row, max_col


def apply_print_settings(ws) -> None:
    max_row, max_col = find_last_used_cell(ws)
    ws.print_area = f"A1:{get_column_letter(max_col)}{max_row}"

    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"

    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    else:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.scale = None
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    ws.page_margins.left = 0.2
    ws.page_margins.right = 0.2
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25
    ws.page_margins.header = 0.1
    ws.page_margins.footer = 0.1

    ws.print_options.horizontalCentered = False
    ws.print_options.verticalCentered = False


def keep_only_target_sheet_visible(workbook, target_ws) -> None:
    for ws in workbook.worksheets:
        if ws is target_ws:
            ws.sheet_state = "visible"
        else:
            ws.sheet_state = "hidden"


def convert_with_libreoffice(
    soffice_bin: str,
    source_file: Path,
    out_dir: Path,
    target_format: str,
    expected_suffix: str,
) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)

    with tempfile.TemporaryDirectory(prefix="soffice_profile_") as profile_dir:
        # Isolated profile avoids profile lock / session issues on headless Pi.
        profile_uri = Path(profile_dir).resolve().as_uri()
        cmd = [
            soffice_bin,
            *SOFFICE_STARTUP_FLAGS,
            f"-env:UserInstallation={profile_uri}",
            "--convert-to",
            target_format,
            "--outdir",
            str(out_dir),
            str(source_file),
        ]

        env = os.environ.copy()
        env.setdefault("SAL_USE_VCLPLUGIN", "svp")

        result = subprocess.run(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            env=env,
        )

    converted_path = out_dir / f"{source_file.stem}{expected_suffix}"

    if result.returncode != 0:
        details = build_libreoffice_error_message(result, source_file)
        raise RuntimeError(details)

    if not converted_path.exists():
        details = build_libreoffice_error_message(
            result,
            source_file,
            fallback=f"{expected_suffix} file was not created",
        )
        raise RuntimeError(details)

    return converted_path


def prepare_excel_for_pdf(path, soffice_bin=None) -> Path:
    src = Path(path).resolve()
    validate_input_file(src)

    if soffice_bin is None:
        soffice_bin = find_soffice()

    if OPENPYXL_IMPORT_ERROR is not None:
        raise ValidationError(
            "Python package 'openpyxl' is required. "
            "Install it with: pip install openpyxl"
        )

    temp_dir = Path(tempfile.mkdtemp(prefix="xlsx2pdf_"))

    source_for_openpyxl = src
    if src.suffix.lower() == ".xls":
        source_for_openpyxl = convert_with_libreoffice(
            soffice_bin=soffice_bin,
            source_file=src,
            out_dir=temp_dir,
            target_format="xlsx",
            expected_suffix=".xlsx",
        )

    prepared_path = temp_dir / f"{source_for_openpyxl.stem}_prepared.xlsx"
    wb = load_workbook(source_for_openpyxl)

    visible_sheets = get_visible_sheets(wb)
    if not visible_sheets:
        raise ValidationError("Workbook has no visible sheets")

    target_ws = visible_sheets[0]

    if is_effectively_empty(target_ws):
        raise ValidationError("First visible worksheet is empty")

    keep_only_target_sheet_visible(wb, target_ws)
    apply_print_settings(target_ws)

    wb.active = wb.worksheets.index(target_ws)

    wb.save(prepared_path)
    return prepared_path


def build_final_pdf_name(prepared_pdf: Path, src: Path, out_dir: Path) -> Path:
    final_pdf = out_dir / f"{src.stem}.pdf"

    if final_pdf.exists():
        final_pdf.unlink()

    prepared_pdf.rename(final_pdf)
    return final_pdf


def convert_word_to_pdf(src: Path, out_dir: Path, soffice_bin: str) -> Path:
    final_pdf = out_dir / f"{src.stem}.pdf"
    if final_pdf.exists():
        final_pdf.unlink()

    return convert_with_libreoffice(
        soffice_bin=soffice_bin,
        source_file=src,
        out_dir=out_dir,
        target_format="pdf",
        expected_suffix=".pdf",
    )


def build_libreoffice_error_message(
    result: subprocess.CompletedProcess,
    source_file: Path,
    fallback: str = "Unknown LibreOffice conversion error",
) -> str:
    stderr = result.stderr.strip()
    stdout = result.stdout.strip()
    details = "\n".join(part for part in (stderr, stdout) if part) or fallback
    low = details.lower()

    hints = []
    if "source file could not be loaded" in low:
        hints.append(
            f"LibreOffice cannot open '{source_file.name}'. "
            "Check file readability and document integrity."
        )
        if source_file.suffix.lower() == ".docx":
            hints.append(
                "On Raspberry Pi install Writer support: "
                "sudo apt install libreoffice-writer"
            )

    if hints:
        details = f"{details}\nHint: {' '.join(hints)}"

    return details


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("input_file")
    parser.add_argument("-o", "--outdir", default="pdf_out")

    args = parser.parse_args()

    src = Path(args.input_file).resolve()
    out_dir = Path(args.outdir).resolve()

    try:
        validate_input_file(src)
        soffice_bin = find_soffice()

        if is_excel_file(src):
            prepared_xlsx = prepare_excel_for_pdf(src, soffice_bin=soffice_bin)
            prepared_pdf = convert_with_libreoffice(
                soffice_bin=soffice_bin,
                source_file=prepared_xlsx,
                out_dir=out_dir,
                target_format="pdf",
                expected_suffix=".pdf",
            )
            final_pdf = build_final_pdf_name(prepared_pdf, src, out_dir)
        else:
            final_pdf = convert_word_to_pdf(src, out_dir, soffice_bin)

        sys.stdout.write(str(final_pdf) + "\n")
        return 0

    except ValidationError as exc:
        sys.stderr.write(f"Validation error: {exc}\n")
        return 2

    except Exception as exc:
        sys.stderr.write(f"Conversion error: {exc}\n")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
